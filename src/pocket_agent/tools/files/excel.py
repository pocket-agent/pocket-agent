from openpyxl import load_workbook

from pocket_agent.config.models import PathsConfig
from pocket_agent.logging.action_log import log_action
from pocket_agent.tools.base import ToolResult
from pocket_agent.tools.files.paths import resolve_read_path
from pocket_agent.tools.files.safety import finalize_safe_edit, prepare_safe_edit
from pocket_agent.tools.files.validators import validate_xlsx


async def analyze_excel(
    paths: PathsConfig,
    file_path: str,
    sample_rows: int = 5,
) -> ToolResult:
    """Read workbook structure and sample cell values. See TOOLS_SPEC analyze_excel()."""
    resolved = resolve_read_path(paths, file_path)
    if isinstance(resolved, ToolResult):
        log_action(
            paths.logs_dir,
            "analyze_excel",
            {"file_path": file_path},
            success=False,
            error=resolved.error,
        )
        return resolved

    path = resolved
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return ToolResult(success=False, error="File is not an Excel workbook (.xlsx)")

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        sheets: list[dict] = []

        for name in wb.sheetnames:
            ws = wb[name]
            rows = ws.max_row or 0
            cols = ws.max_column or 0
            sample: list[list[str]] = []
            for row in ws.iter_rows(min_row=1, max_row=min(sample_rows, rows), values_only=True):
                sample.append([str(c) if c is not None else "" for c in row])

            sheets.append(
                {
                    "name": name,
                    "rows": rows,
                    "columns": cols,
                    "sample": sample,
                }
            )

        wb.close()

        log_action(
            paths.logs_dir,
            "analyze_excel",
            {"file_path": str(path), "sheet_count": len(sheets)},
        )
        return ToolResult(
            success=True,
            data={"path": str(path), "sheets": sheets, "sheet_count": len(sheets)},
        )
    except Exception as exc:
        log_action(
            paths.logs_dir,
            "analyze_excel",
            {"file_path": str(path)},
            success=False,
            error=str(exc),
        )
        return ToolResult(success=False, error=str(exc))


async def modify_excel(
    paths: PathsConfig,
    file_path: str,
    sheet_name: str,
    cell: str,
    value: str,
) -> ToolResult:
    """Apply a single cell change with backup, validation, and replace. See TOOLS_SPEC."""
    resolved = resolve_read_path(paths, file_path)
    if isinstance(resolved, ToolResult):
        log_action(
            paths.logs_dir,
            "modify_excel",
            {"file_path": file_path},
            success=False,
            error=resolved.error,
        )
        return resolved

    path = resolved
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return ToolResult(success=False, error="File is not an Excel workbook (.xlsx)")

    session = prepare_safe_edit(paths, path, "modify_excel")
    if isinstance(session, ToolResult):
        return session

    try:
        wb = load_workbook(session.working_path)
        if sheet_name not in wb.sheetnames:
            return ToolResult(success=False, error=f"Sheet not found: {sheet_name}")
        ws = wb[sheet_name]
        ws[cell] = _coerce_cell_value(value)
        wb.save(session.working_path)
        wb.close()
    except Exception as exc:
        log_action(
            paths.logs_dir,
            "modify_excel",
            {"file_path": str(path), "cell": cell},
            success=False,
            error=str(exc),
        )
        return ToolResult(success=False, error=str(exc))

    result = finalize_safe_edit(paths, session, validate_xlsx, "modify_excel")
    if result.success:
        result.data["sheet"] = sheet_name
        result.data["cell"] = cell
        result.data["value"] = value
    return result


def _coerce_cell_value(value: str) -> str | int | float:
    if value.lower() in {"true", "false"}:
        return value.lower() == "true"
    try:
        if "." in value:
            return float(value)
        return int(value)
    except ValueError:
        return value
