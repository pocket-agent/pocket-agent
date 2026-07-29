from pathlib import Path

import fitz
from docx import Document
from openpyxl import load_workbook


def validate_xlsx(path: Path) -> tuple[bool, str]:
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return False, "not an xlsx file"
    try:
        wb = load_workbook(path, read_only=True)
        if not wb.sheetnames:
            wb.close()
            return False, "workbook has no sheets"
        wb.close()
        return True, ""
    except Exception as exc:
        return False, str(exc)


def validate_pdf(path: Path) -> tuple[bool, str]:
    if path.suffix.lower() != ".pdf":
        return False, "not a pdf file"
    try:
        with fitz.open(path) as doc:
            if doc.page_count < 1:
                return False, "pdf has no pages"
        return True, ""
    except Exception as exc:
        return False, str(exc)


def validate_docx(path: Path) -> tuple[bool, str]:
    if path.suffix.lower() != ".docx":
        return False, "not a docx file"
    try:
        doc = Document(path)
        if doc.element.body is None:
            return False, "invalid document body"
        return True, ""
    except Exception as exc:
        return False, str(exc)
