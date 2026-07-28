from pathlib import Path

import fitz
import pytest
from docx import Document
from openpyxl import Workbook

from pocket_agent.config.models import PathsConfig
from pocket_agent.tools.files.excel import analyze_excel
from pocket_agent.tools.files.pdf import extract_pdf_text
from pocket_agent.tools.files.read import read_file


def _paths(tmp_path: Path) -> PathsConfig:
    nas = tmp_path / "nas"
    nas.mkdir()
    return PathsConfig(
        {
            "nas": {"root": str(nas), "allowed_read_roots": [str(nas)]},
            "data": {"logs": "data/logs"},
        },
        tmp_path,
    )


def _make_txt(nas: Path) -> Path:
    path = nas / "note.txt"
    path.write_text("hello pocket agent")
    return path


def _make_docx(nas: Path) -> Path:
    path = nas / "doc.docx"
    doc = Document()
    doc.add_paragraph("hello docx")
    doc.save(path)
    return path


def _make_xlsx(nas: Path) -> Path:
    path = nas / "sheet.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget"
    ws["A1"] = "Item"
    ws["B1"] = "Amount"
    ws["A2"] = "Rent"
    ws["B2"] = 1000
    wb.save(path)
    return path


def _make_pdf(nas: Path) -> Path:
    path = nas / "sample.pdf"
    doc = fitz.open()
    page = doc.new_page()
    page.insert_text((72, 72), "hello pdf text")
    doc.save(path)
    doc.close()
    return path


@pytest.mark.asyncio
async def test_read_txt(tmp_path: Path):
    paths = _paths(tmp_path)
    paths.logs_dir.mkdir(parents=True, exist_ok=True)
    txt = _make_txt(paths.nas_root)
    result = await read_file(paths, str(txt))
    assert result.success
    assert "hello pocket" in result.data["text"]


@pytest.mark.asyncio
async def test_read_docx(tmp_path: Path):
    paths = _paths(tmp_path)
    paths.logs_dir.mkdir(parents=True, exist_ok=True)
    docx = _make_docx(paths.nas_root)
    result = await read_file(paths, str(docx))
    assert result.success
    assert "hello docx" in result.data["text"]


@pytest.mark.asyncio
async def test_extract_pdf_text(tmp_path: Path):
    paths = _paths(tmp_path)
    paths.logs_dir.mkdir(parents=True, exist_ok=True)
    pdf = _make_pdf(paths.nas_root)
    result = await extract_pdf_text(paths, str(pdf))
    assert result.success
    assert "hello pdf" in result.data["text"]


@pytest.mark.asyncio
async def test_analyze_excel(tmp_path: Path):
    paths = _paths(tmp_path)
    paths.logs_dir.mkdir(parents=True, exist_ok=True)
    xlsx = _make_xlsx(paths.nas_root)
    result = await analyze_excel(paths, str(xlsx))
    assert result.success
    assert result.data["sheet_count"] == 1
    assert result.data["sheets"][0]["name"] == "Budget"
